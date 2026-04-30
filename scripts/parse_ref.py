#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 ref.xlsx 转换为行业分类 JSON 树结构

数据层级：
- 门类 (A-Z, 列0)
- 大类 (2位数字, 列0)
- 中类 (3位数字, 列0) - 包括728这样的中类
- 细类 (4位数字, 列1)
- ◇行 (列2="◇", 列3=标题)
- ◆行 (列2="◆", 列3=标题)
- —子项行 (列3="—", 列4=内容)
"""

import json
from openpyxl import load_workbook

def is_single_letter(s):
    """判断是否为单个字母（门类代码）"""
    if s is None:
        return False
    s = str(s)
    return len(s) == 1 and s.isalpha()

def is_n_digits(s, n):
    """判断是否为n位数字字符串"""
    if s is None:
        return False
    s = str(s)
    return len(s) == n and s.isdigit()

def get_category_letter(digit):
    """根据大类代码第一位数字获取门类字母"""
    digit_map = {'0': 'A', '1': 'B', '2': 'C', '3': 'D', '4': 'E',
                 '5': 'F', '6': 'G', '7': 'H', '8': 'I', '9': 'J'}
    return digit_map.get(digit, '')

def build_tree(input_file, output_file):
    """构建行业分类树"""
    wb = load_workbook(input_file, data_only=True)
    ws = wb.active

    # 根节点
    root = {"code": "root", "name": "国民经济行业分类", "level": 0, "children": [], "description": ""}

    # 各层级节点存储
    level1_map = {}  # 门类字母 -> node
    level2_map = {}  # 大类(2位) -> node
    level3_map = {}  # 中类(3位) -> node

    # 当前活跃的节点和描述行
    current_node = None
    current_detail_lines = []
    in_bukua_section = False  # 是否在"不包括"部分
    current_category = None   # 当前门类上下文（用于大类定位）

    for row in ws.iter_rows(min_row=1, values_only=True):
        col0 = row[0] if len(row) > 0 else None  # A列
        col1 = row[1] if len(row) > 1 else None  # B列
        col2 = row[2] if len(row) > 2 else None  # C列
        col3 = row[3] if len(row) > 3 else None  # D列
        col4 = row[4] if len(row) > 4 else None  # E列

        # 跳过标题行和空行
        if col0 == "2017年国民经济行业分类注释":
            continue
        if col0 is None and col1 is None and col3 is None:
            continue

        # 1. 门类节点 (单个字母)
        if is_single_letter(col0):
            # 保存上一个节点的描述
            if current_node is not None and current_detail_lines:
                current_node["description"] = "\n".join(current_detail_lines)

            node = {
                "code": col0,
                "name": str(col3) if col3 else "",
                "level": 1,
                "children": [],
                "description": ""
            }
            root["children"].append(node)
            level1_map[col0] = node
            current_node = node
            current_detail_lines = []
            in_bukua_section = False
            current_category = col0  # 更新当前门类上下文
            continue

        # 2. 大类节点 (2位数字)
        if is_n_digits(col0, 2):
            # 保存上一个节点的描述
            if current_node is not None and current_detail_lines:
                current_node["description"] = "\n".join(current_detail_lines)

            col0_str = str(col0)
            parent = level1_map.get(current_category)  # 使用当前门类上下文
            node = {
                "code": col0_str,
                "name": str(col3) if col3 else "",
                "level": 2,
                "children": [],
                "description": ""
            }
            if parent:
                parent["children"].append(node)
            level2_map[col0_str] = node
            current_node = node
            current_detail_lines = []
            in_bukua_section = False
            continue

        # 3. 同一行有中类(3位)和细类(4位)时，先处理中类
        has_mid_class = is_n_digits(col0, 3)
        has_leaf_class = is_n_digits(col1, 4)

        # 保存上一个节点的描述
        if current_node is not None and current_detail_lines:
            current_node["description"] = "\n".join(current_detail_lines)

        # 中类节点 (3位数字) - 先处理
        if has_mid_class:
            col0_str = str(col0)
            parent = level2_map.get(col0_str[:2])
            node = {
                "code": col0_str,
                "name": str(col3) if col3 else "",
                "level": 3,
                "children": [],
                "description": ""
            }
            if parent:
                parent["children"].append(node)
            level3_map[col0_str] = node
            current_node = node
            current_detail_lines = []
            in_bukua_section = False

        # 细类节点 (4位数字在B列) - 后处理，确保父节点已创建
        if has_leaf_class:
            col1_str = str(col1)
            # 找父节点（优先3位，再2位）
            parent = level3_map.get(col1_str[:3])
            if parent is None:
                parent = level2_map.get(col1_str[:2])

            node = {
                "code": col1_str,
                "name": str(col3) if col3 else "",
                "level": 4,
                "children": [],
                "description": ""
            }
            if parent:
                parent["children"].append(node)
            current_node = node
            current_detail_lines = []
            in_bukua_section = False

        # 如果处理了中类或细类，跳过后续处理
        if has_mid_class or has_leaf_class:
            continue

        # 5. ◇包括标题行
        if col2 == "◇":
            if current_node is not None and col3:
                current_detail_lines.append("◇ " + str(col3))
                in_bukua_section = False
            continue

        # 6. ◆不包括标题行
        if col2 == "◆":
            if current_node is not None and col3:
                current_detail_lines.append("◆ " + str(col3))
                in_bukua_section = True
            continue

        # 7. —子项行（所有内容都用"——"，不管是"包括"还是"不包括"部分）
        if col3 == "—":
            if current_node is not None and col4:
                line = str(col4).rstrip("；").rstrip("。")
                current_detail_lines.append("— " + line)
            continue

        # 8. 直接跟在节点后面的描述行
        # 跳过第一行废话（如"◇ 指xxx"），只保留"◇ 包括..."或"◇ 包括下列..."
        if col0 is None and col1 is None and col2 is None and current_node is not None:
            if col3 and str(col3).strip():
                text = str(col3).strip()
                # 不是"—"且不是节点名称
                if text != "—" and text != current_node["name"]:
                    # 跳过"不包括"开头的行（将在◇包括之后出现）
                    if text.startswith("不包括"):
                        continue
                    # 跳过"指..."这种定义性描述，只保留"包括..."开头的行
                    if "包括" in text:
                        current_detail_lines.append("◇ " + text)
            continue

    # 保存最后一个节点的描述
    if current_node is not None and current_detail_lines:
        current_node["description"] = "\n".join(current_detail_lines)

    # 统计
    all_nodes = []
    def flatten(node):
        all_nodes.append(node)
        for child in node.get("children", []):
            flatten(child)

    flatten(root)

    level_counts = {}
    for node in all_nodes:
        level = node["level"]
        level_counts[level] = level_counts.get(level, 0) + 1

    leaf_with_desc = sum(1 for node in all_nodes if node["level"] == 4 and node["description"])

    # 写入JSON
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(root, f, ensure_ascii=False, indent=2)

    print(f"=== 解析完成 ===")
    print(f"总节点数: {len(all_nodes)}")
    print(f"层级分布: {level_counts}")
    print(f"细类(4位)数量: {level_counts.get(4, 0)}")
    print(f"有description的细类: {leaf_with_desc}")
    print(f"输出文件: {output_file}")

    return root

if __name__ == "__main__":
    input_file = "ref.xlsx"
    output_file = "data/industry_tree.json"
    build_tree(input_file, output_file)
