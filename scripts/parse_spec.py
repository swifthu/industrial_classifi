#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
解析 spec.xlsx 生成增强版行业分类数据
从行业分类标准Excel中提取高技术、知识产权密集型、战略性新兴产业等标签信息
"""

import os
import json
import re
import openpyxl

# ==================== 模块级配置常量 ====================
WORK_DIR = "/Users/jimmyhu/Documents/Python/industrial_classifi"
LEVEL4_CODE_LENGTH = 4
SPEC_FILE = os.path.join(WORK_DIR, "spec.xlsx")
BASE_DATA_FILE = os.path.join(WORK_DIR, "data", "industry_tree_basic.json")
OUTPUT_FILE = os.path.join(WORK_DIR, "data", "industry_tree_advanced.json")

# 标签定义
TAGS_MAP = {
    "highTech_mfg": "高技术（制造业）",
    "highTech_svc": "高技术（服务业）",
    "ip密集型": "知识产权密集型产业",
    "strategic": "战略性新兴产业",
    "digital": "数字经济核心产业",
    "pension": "养老产业",
    "culture": "文化产业"
}

# 各Sheet配置: (tag_key, 代码列索引, 说明列索引, 数据起始行)
# 每行列含义：
# 高技术制造: A(大类),B(中类),C(小类),D(名称),E(代码),F(代码2),G(分类)
# 高技术服务: A(大类),B(中类),C(小类),D(名称),E(代码),F(说明),G(代码2),H(分类)
# 知识产权: A(大类),B(中类),C(分类名),D(国民经济代码),E(名称),F(代码),G(分类)
# 战略性: A(分类代码),B(分类名),C(代码),D(行业名),E(代码2),F(分类)
# 数字经济: A(大类),B(中类),C(小类),D(名称),E(说明),F(代码),G(代码2),H(分类)
# 养老: A(大类),B(中类),C(小类),D(名称),E(说明),F(代码),I(分类)
# 文化: A(大类),B(中类),C(小类),D(名称),E(说明),F(代码),G(分类)
SHEET_CONFIG = {
    "高技术（制造业）": ("highTech_mfg", 4, None, 3),
    "高技术（服务业）": ("highTech_svc", 4, 5, 4),
    "知识产权密集型产业": ("ip密集型", 5, None, 6),
    "战略性新兴产业 ": ("strategic", 2, 5, 4),
    "数字经济核心产业": ("digital", 5, 4, 6),
    "养老产业": ("pension", 5, 4, 5),
    "文化产业": ("culture", 5, 4, 8),
}

STRATEGIC_PRODUCTS_SHEET = "战略新兴产业重点说明"
STRATEGIC_PRODUCTS_CODE_COL = 2
STRATEGIC_PRODUCTS_COL = 4


def load_base_data():
    try:
        with open(BASE_DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"基础数据文件不存在: {BASE_DATA_FILE}")
    except json.JSONDecodeError as e:
        raise json.JSONDecodeError(f"JSON解析失败: {e}", doc=str(e), pos=0)


def collect_level4_nodes(node, nodes_dict):
    if node["level"] == 4:
        nodes_dict[node["code"]] = node
    for child in node.get("children", []):
        collect_level4_nodes(child, nodes_dict)


def extract_codes_from_string(code_str):
    results = []
    if not code_str:
        return results

    pattern = rf'(\d{{{LEVEL4_CODE_LENGTH}}})(\*?)'
    matches = re.findall(pattern, code_str)
    for match in matches:
        code_4d = match[0]
        has_star = match[1] == '*'
        original = code_4d + ('*' if has_star else '')
        results.append((code_4d, has_star, original))

    return results


def build_strategic_products_mapping():
    wb = openpyxl.load_workbook(SPEC_FILE, data_only=True)

    if STRATEGIC_PRODUCTS_SHEET not in wb.sheetnames:
        print(f"警告: Sheet '{STRATEGIC_PRODUCTS_SHEET}' 不存在")
        wb.close()
        return {}

    ws = wb[STRATEGIC_PRODUCTS_SHEET]
    code_products = {}
    current_code = None

    for row in ws.iter_rows(min_row=4, values_only=True):
        col_c = row[STRATEGIC_PRODUCTS_CODE_COL]
        col_e = row[STRATEGIC_PRODUCTS_COL]

        if col_c and str(col_c).strip():
            code_str = str(col_c).strip()
            if code_str.endswith('*'):
                current_code = code_str
                code_products[current_code] = []

        if current_code and col_e and str(col_e).strip():
            code_products[current_code].append(str(col_e).strip())

    wb.close()
    return code_products


def update_category_path(path, a, b, c, d, is_strategic=False, large_cat=None):
    """
    根据行列信息更新category路径

    普通sheet: A(大类),B(中类),C(小类),D(名称)
    战略性sheet: A(分类代码如1.1.3),B(分类名称),C(代码),D(国民经济行业名称),large_cat(F列大类名)
    """
    if path is None:
        path = []

    b_str = str(b).strip() if b else ""
    c_str = str(c).strip() if c else ""
    d_str = str(d).strip() if d else ""

    if is_strategic:
        # 战略性新兴产业：解析A列的分类代码（如1.1.3）来构建路径
        # A列格式：大类.中类.小类，或只有中类如1.1，或只有大类如1
        if a is not None and b_str:
            a_str = str(a).strip()
            parts = a_str.split('.')
            level = len(parts)  # 1=大类, 2=中类, 3=小类

            if level == 1:
                # 大类，如"1" -> "新一代信息技术产业"
                path = [b_str]
            elif level == 2:
                # 中类，如"1.1" -> "下一代信息网络产业"
                if large_cat:
                    path = [large_cat, b_str]
                else:
                    path = [b_str]
            elif level >= 3:
                # 小类，如"1.1.3" -> 设置为第三级，追加而不是覆盖
                if large_cat:
                    if len(path) >= 2 and path[0] == large_cat:
                        # 大类已存在，替换第三级
                        path = [path[0], path[1] if len(path) > 1 else '', b_str]
                    else:
                        # 新的大类
                        path = [large_cat, b_str]
                else:
                    path = [b_str]
            return path
        elif a is None and b_str == '' and d_str:
            # 叶子行（A=None），使用当前path，追加小类名
            path = path + [d_str]
            return path
        return path if path else []
    else:
        # B有值 = 中类（B列是代码，说明有子分类）
        if b is not None and d_str:
            # c有值说明B是中类代码，C是中类名称
            if c is not None:
                if path and len(path) >= 1 and path[0]:
                    return [path[0], d_str, ""]
                return [d_str, ""]
            else:  # 只有两级（没有小类）
                if path and len(path) >= 1 and path[0]:
                    return [path[0], d_str]
                return [d_str]
        # C有值 = 小类（叶子节点）
        if c is not None and d_str:
            if path:
                if len(path) >= 2 and path[-1] == '':
                    return path[:-1] + [d_str]
                elif len(path) >= 2:
                    return path[:2] + [d_str]
                return path + [d_str]
            return [d_str]
        # A有值但B/C为空 = 新的大类开始，重置path
        if a is not None and d_str and b is None and c is None:
            return [d_str]
        # A/B/C都为空但D有时（中间分类名）
        if d_str and a is None and b is None and c is None:
            if path:
                if len(path) >= 1 and path[-1] == '':
                    # 替换空字符串
                    return path[:-1] + [d_str]
                return path + [d_str]
            return [d_str]
        return path if path else []


def build_tag_mapping():
    wb = openpyxl.load_workbook(SPEC_FILE, data_only=True)

    strategic_products = build_strategic_products_mapping()
    print(f"      战略性新兴产业产品目录包含 {len(strategic_products)} 个代码")

    tag_mapping = {}

    for sheet_name, (tag_key, code_col_idx, desc_col_idx, min_row) in SHEET_CONFIG.items():
        matched_name = None
        for ws_name in wb.sheetnames:
            if ws_name.strip() == sheet_name.strip() or ws_name == sheet_name:
                matched_name = ws_name
                break

        if matched_name is None:
            print(f"警告: Sheet '{sheet_name}' 不存在，跳过")
            continue

        ws = wb[matched_name]
        category_path = []
        is_strategic = (tag_key == "strategic")
        is_ip = (tag_key == "ip密集型")

        for row in ws.iter_rows(min_row=min_row, values_only=True):
            col_a = row[0]
            col_b = row[1]
            col_c = row[2]
            col_d = row[3]

            # 知识产权sheet用不同的列构建路径
            if is_ip:
                col_g = row[6] if len(row) > 6 else None  # G列是大类名

                if col_a is not None and col_c:
                    # A列有值时表示新的大类开始，path需要重置
                    category_path = []
                    large_cat = col_c
                    name_col = col_c  # 大类名
                elif col_g:
                    # G列有值，表示继承当前大类
                    large_cat = col_g
                    if col_b is not None and col_c:
                        # B有值=中类行，name_col用C列（中类名），leaf用E列（具体名称）
                        name_col = col_c  # 中类名
                    elif row[4] and not str(row[4]).isdigit():
                        # B为空但E列有非数字名称，用E列
                        name_col = row[4]
                    else:
                        name_col = col_c if col_c else ''
                else:
                    large_cat = None
                    name_col = ''
            else:
                name_col = col_d  # D列是名称
                # 对于战略性sheet，F列是大类名
                large_cat = row[5] if is_strategic and len(row) > 5 else col_a

            # 更新路径
            # IP密集型sheet结构:
            # - 大类行: A有值, C是大类名
            # - 中类行: B有值, C是中类名, F有代码
            # - 叶子行: B为空, F有代码, E是具体名称
            if is_ip:
                if col_a is not None and col_c:
                    # 大类行
                    category_path = [large_cat]
                elif col_b is not None and col_c:
                    # 中类行：B有值，创建[大类, 中类名, '']
                    category_path = [large_cat, col_c.strip(), '']
                elif col_b is None and col_c is None and large_cat and name_col:
                    # 叶子行: B和C都为空，用叶子名称替换path的最后一个元素
                    if len(category_path) >= 1:
                        category_path[-1] = name_col
                    else:
                        category_path = [name_col]
                else:
                    category_path = update_category_path(category_path, large_cat, col_b, col_c, name_col, is_strategic)
            elif is_strategic:
                # 战略性sheet结构: A=分类代码(如1.1.3), B=分类名称, C=代码, D=行业名称, F=大类名
                if col_a is not None and col_b:
                    a_str = str(col_a).strip()
                    parts = a_str.split('.')
                    level = len(parts)

                    current_large = large_cat if large_cat else ''

                    if level == 1:
                        # 大类
                        category_path = [col_b]
                        if 'mid_cache' not in tag_mapping:
                            tag_mapping['mid_cache'] = {}
                    elif level == 2:
                        # 中类
                        mid_cache = tag_mapping.get('mid_cache', {})
                        mid_cache[a_str] = col_b
                        tag_mapping['mid_cache'] = mid_cache
                        category_path = [current_large, col_b] if current_large else [col_b]
                    elif level >= 3:
                        # 小类
                        mid_cache = tag_mapping.get('mid_cache', {})
                        mid_code = '.'.join(parts[:-1])
                        mid_name = mid_cache.get(mid_code, '')
                        category_path = [current_large, mid_name, col_b] if current_large else [mid_name, col_b]
                elif col_a is None and col_b is None and col_c is None and col_d:
                    # 叶子行: C为空，D是行业名称，追加到path
                    category_path = category_path + [col_d]
                # 如果C有值（代码行），不修改path，让代码使用当前path
            else:
                category_path = update_category_path(category_path, large_cat, col_b, col_c, name_col, is_strategic)

            # 获取代码
            code_cell = row[code_col_idx]
            if not code_cell:
                continue

            code_str = str(code_cell).strip()
            codes_info = extract_codes_from_string(code_str)

            for code_4d, has_star, original_code in codes_info:
                description = ""
                if has_star and desc_col_idx is not None:
                    desc_cell = row[desc_col_idx]
                    if desc_cell and str(desc_cell).strip():
                        description = str(desc_cell).strip()

                # 对于战略性新兴产业，同一个代码可能在不同分类中出现多次，需要收集所有分类
                if tag_key == "strategic" and code_4d in tag_mapping and "strategic" in tag_mapping[code_4d]:
                    # 已有记录，收集多个分类
                    existing = tag_mapping[code_4d]["strategic"]
                    if "categories" not in existing:
                        existing["categories"] = [existing.get("category", "")] if existing.get("category") else []
                    cat = " > ".join(filter(None, category_path))
                    if cat and cat not in existing["categories"]:
                        existing["categories"].append(cat)
                    # description只保留第一个有description的
                    if description and not existing.get("description"):
                        existing["description"] = description
                    continue

                if code_4d not in tag_mapping:
                    tag_mapping[code_4d] = {}

                if has_star:
                    detail = {
                        "has_star": True,
                        "category": " > ".join(filter(None, category_path)),
                        "description": description if description else ""
                    }
                    if tag_key == "strategic":
                        products = strategic_products.get(original_code, [])
                        if products:
                            detail["products"] = products
                    tag_mapping[code_4d][tag_key] = detail
                else:
                    tag_mapping[code_4d][tag_key] = {
                        "has_star": False,
                        "category": " > ".join(filter(None, category_path)),
                        "description": ""
                    }

    wb.close()
    return tag_mapping


def enhance_data():
    print("=" * 50)
    print("开始解析 spec.xlsx 生成增强版数据")
    print("=" * 50)

    print("\n[1/4] 加载基础数据...")
    base_data = load_base_data()

    print("[2/4] 收集四级节点...")
    level4_nodes = {}
    collect_level4_nodes(base_data, level4_nodes)
    print(f"      找到 {len(level4_nodes)} 个四级节点")

    print("[3/4] 从 spec.xlsx 构建标签映射...")
    tag_mapping = build_tag_mapping()
    print(f"      映射表中包含 {len(tag_mapping)} 个代码")

    print("[4/4] 增强四级节点数据...")

    stats = {
        "total": len(level4_nodes),
        "tagged": 0,
        "by_tag": {tag: 0 for tag in TAGS_MAP.keys()}
    }

    def enhance_node(node):
        if node["level"] == 4:
            code = node["code"]
            node["tags"] = []
            node["tagDetail"] = {}

            code_prefix = code[:LEVEL4_CODE_LENGTH]

            if code_prefix in tag_mapping:
                for tag_key, tag_info in tag_mapping[code_prefix].items():
                    if tag_key not in node["tags"]:
                        node["tags"].append(tag_key)
                        stats["tagged"] += 1
                        stats["by_tag"][tag_key] += 1
                    node["tagDetail"][tag_key] = tag_info

        for child in node.get("children", []):
            enhance_node(child)

    enhance_node(base_data)

    print("\n[保存] 输出到 industry_tree_advanced.json...")
    try:
        with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
            json.dump(base_data, f, ensure_ascii=False, indent=2)
    except IOError as e:
        raise IOError(f"写入输出文件失败: {OUTPUT_FILE}, 错误: {e}")

    print("\n" + "=" * 50)
    print("处理完成!")
    print("=" * 50)
    print(f"四级节点总数: {stats['total']}")
    print(f"带标签的节点: {stats['tagged']}")
    print("\n各标签分布:")
    for tag_key, count in stats["by_tag"].items():
        tag_name = TAGS_MAP.get(tag_key, tag_key)
        print(f"  - {tag_name}: {count}")

    print(f"\n输出文件: {OUTPUT_FILE}")

    return stats


if __name__ == "__main__":
    enhance_data()